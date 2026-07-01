#!/usr/bin/env python3
"""
Export ranked profiles to CSV and XLSX formats
- CSV: Individual files for each role with all 50 ranked profiles
- XLSX: Consolidated file with one sheet per role
"""

import json
import csv
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuration
OUTPUT_DIR = '/home/anand/projects/candidate_ranking/Ideathon/team_hiring_solution/source/outputs'
JSON_FILE = os.path.join(OUTPUT_DIR, 'team_hiring_report_20260701_113828.json')

def load_json_data(filepath):
    """Load ranking data from JSON file"""
    with open(filepath, 'r') as f:
        return json.load(f)

def export_to_csv(data, output_dir):
    """Export each role's rankings to individual CSV files"""
    print("\n📄 Creating CSV files per role...")
    csv_files = []
    
    for role_id, role_data in data['per_role_analysis'].items():
        all_candidates = role_data.get('all_ranked_candidates', [])
        
        # Create CSV filename
        csv_filename = f"{role_id}_ranked_profiles.csv"
        csv_filepath = os.path.join(output_dir, csv_filename)
        
        # Write CSV file
        with open(csv_filepath, 'w', newline='') as f:
            writer = csv.writer(f)
            
            # Header
            writer.writerow(['Rank', 'Candidate ID', 'Score', 'Skill Coverage'])
            
            # Data rows
            for profile in all_candidates:
                writer.writerow([
                    profile['rank'],
                    profile['candidate_id'],
                    f"{profile['score']:.4f}",
                    profile['skill_coverage']
                ])
        
        csv_files.append(csv_filepath)
        print(f"   ✅ {csv_filename} - {len(all_candidates)} profiles")
    
    return csv_files

def export_to_xlsx(data, output_dir):
    """Export all roles to XLSX with one sheet per role"""
    print("\n📊 Creating consolidated XLSX file...")
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Alternating row colors
    light_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    role_count = 0
    
    for role_id, role_data in data['per_role_analysis'].items():
        all_candidates = role_data.get('all_ranked_candidates', [])
        
        # Create sheet for this role
        ws = wb.create_sheet(title=role_id.replace('_', ' ')[:31])  # Sheet name max 31 chars
        
        # Add header
        headers = ['Rank', 'Candidate ID', 'Score', 'Skill Coverage']
        ws.append(headers)
        
        # Format header
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data rows
        for row_num, profile in enumerate(all_candidates, 2):
            ws.cell(row=row_num, column=1, value=profile['rank'])
            ws.cell(row=row_num, column=2, value=profile['candidate_id'])
            ws.cell(row=row_num, column=3, value=round(profile['score'], 4))
            ws.cell(row=row_num, column=4, value=profile['skill_coverage'])
            
            # Format data row
            for col_num in range(1, 5):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = data_alignment
                cell.border = border
                
                # Alternating row colors
                if row_num % 2 == 0:
                    cell.fill = light_fill
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        role_count += 1
        print(f"   ✅ Sheet: {role_id} - {len(all_candidates)} profiles")
    
    # Save XLSX file
    xlsx_filename = 'ranked_profiles_consolidated.xlsx'
    xlsx_filepath = os.path.join(output_dir, xlsx_filename)
    wb.save(xlsx_filepath)
    
    print(f"\n   ✅ Saved: {xlsx_filename} ({role_count} sheets)")
    return xlsx_filepath

def create_summary_report(output_dir, csv_files, xlsx_filepath):
    """Create a summary of exported files"""
    summary_file = os.path.join(output_dir, 'EXPORT_SUMMARY.txt')
    
    with open(summary_file, 'w') as f:
        f.write("=" * 80 + "\n")
        f.write("RANKED PROFILES EXPORT SUMMARY\n")
        f.write("=" * 80 + "\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        
        f.write("📄 CSV FILES (Per Role)\n")
        f.write("-" * 80 + "\n")
        for csv_file in csv_files:
            filename = os.path.basename(csv_file)
            filesize = os.path.getsize(csv_file)
            f.write(f"  ✅ {filename} ({filesize:,} bytes)\n")
        
        f.write("\n📊 XLSX FILE (Consolidated)\n")
        f.write("-" * 80 + "\n")
        filesize = os.path.getsize(xlsx_filepath)
        f.write(f"  ✅ {os.path.basename(xlsx_filepath)} ({filesize:,} bytes)\n")
        f.write(f"     Sheets: 5 (one per role)\n")
        f.write(f"     Rows per sheet: 51 (1 header + 50 profiles)\n")
        
        f.write("\n" + "=" * 80 + "\n")
        f.write("USAGE\n")
        f.write("=" * 80 + "\n")
        f.write("CSV Files:\n")
        f.write("  - Open individual role files in any spreadsheet application\n")
        f.write("  - Each file contains all 50 ranked profiles with scores\n")
        f.write("  - Ideal for filtering, sorting, or importing into HR systems\n\n")
        f.write("XLSX File:\n")
        f.write("  - Open consolidated file in Excel/Sheets\n")
        f.write("  - Switch tabs to view different roles\n")
        f.write("  - Formatted with headers, alternating colors, borders\n")
        f.write("  - Frozen header row for easy scrolling\n\n")
        f.write("=" * 80 + "\n")
    
    return summary_file

def main():
    print("=" * 80)
    print("RANKED PROFILES EXPORT TOOL")
    print("=" * 80)
    
    # Load JSON data
    print(f"\n📖 Loading data from: {JSON_FILE}")
    data = load_json_data(JSON_FILE)
    
    # Export to CSV
    csv_files = export_to_csv(data, OUTPUT_DIR)
    
    # Export to XLSX
    xlsx_filepath = export_to_xlsx(data, OUTPUT_DIR)
    
    # Create summary
    summary_file = create_summary_report(OUTPUT_DIR, csv_files, xlsx_filepath)
    
    # Final summary
    print("\n" + "=" * 80)
    print("✅ EXPORT COMPLETE!")
    print("=" * 80)
    print(f"\n📁 Output Directory: {OUTPUT_DIR}\n")
    
    print("📄 CSV Files (5 files):")
    for csv_file in csv_files:
        filename = os.path.basename(csv_file)
        print(f"   ✅ {filename}")
    
    print(f"\n📊 XLSX File (1 file):")
    print(f"   ✅ {os.path.basename(xlsx_filepath)}")
    
    print(f"\n📋 Summary: {os.path.basename(summary_file)}")
    
    print("\n" + "=" * 80)
    print("Files are ready to download and use! 🎉")
    print("=" * 80)

if __name__ == '__main__':
    main()
